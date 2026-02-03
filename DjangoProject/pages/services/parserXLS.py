from pathlib import Path
from typing import Any, Dict, List
from datetime import datetime
import openpyxl
import shutil


from pathlib import Path
from typing import Dict, List, Any
import openpyxl
import shutil
from datetime import datetime


def load_session_from_excel(
    session_id: int,
    file_path: str,
    move_done: bool = False,
) -> Dict[int, List[Any]]:
    """
    Читает Excel, выбранный пользователем, и возвращает данные
    session_id — id сессии отбора
    file_path  — ПОЛНЫЙ путь к Excel (из IHFileSelect)
    move_done  — переносить ли файл в ComplenXLS после обработки
    """

    xls_path = Path(file_path)
    if not xls_path.exists():
        raise FileNotFoundError(f"Excel-файл не найден: {xls_path}")

    excel_exts = {".xlsx", ".xlsm", ".xltx", ".xltm"}
    if xls_path.suffix.lower() not in excel_exts:
        raise ValueError(f"Файл {xls_path.name} не является Excel")

    # 1) читаем Excel
    wb = openpyxl.load_workbook(xls_path, data_only=True)
    ws = wb.active  # один лист

    # ─── настройки под твой формат ──────────────────────────────
    key_col = 2              # B — проверка пустых строк
    read_cols = (2, 3, 6)    # B, C, F
    start_row = 11
    empty_limit = 3
    # ────────────────────────────────────────────────────────────

    result: Dict[int, List[Any]] = {}
    idx = 1
    empty_streak = 0
    row = start_row

    while True:
        key = ws.cell(row=row, column=key_col).value

        if key is None or (isinstance(key, str) and key.strip() == ""):
            empty_streak += 1
            if empty_streak >= empty_limit:
                break
            row += 1
            continue

        empty_streak = 0

        values: List[Any] = []
        for c in read_cols:
            v = ws.cell(row=row, column=c).value
            v = v.strip() if isinstance(v, str) else v
            values.append(v)

        result[idx] = values
        idx += 1
        row += 1

    wb.close()

    # 2) перенос файла (если нужно)
    if move_done:
        done_dir = Path("ComplenXLS")
        done_dir.mkdir(parents=True, exist_ok=True)

        today = datetime.now().strftime("%Y-%m-%d")
        new_name = f"{today}_session_{session_id}{xls_path.suffix}"
        dest_path = done_dir / new_name

        shutil.move(str(xls_path), dest_path)

    return result



