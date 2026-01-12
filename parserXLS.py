from pathlib import Path
from typing import Any, Dict, List
from datetime import datetime
import openpyxl
import shutil


def load_session_from_excel(session_id: int, folder: str = "ExcelSource") -> Dict[int, List[Any]]:
    src_dir = Path(folder)
    done_dir = Path("ComplenXLS")
    src_dir.mkdir(parents=True, exist_ok=True)
    done_dir.mkdir(parents=True, exist_ok=True)

    excel_exts = {".xlsx", ".xlsm", ".xltx", ".xltm"}

    # 1) берем последний Excel-файл в папке (без привязки к имени)
    files = [p for p in src_dir.iterdir() if p.is_file() and p.suffix.lower() in excel_exts]
    if not files:
        raise FileNotFoundError(f"В папке {src_dir} не найдено Excel-файлов")

    xls_path = max(files, key=lambda x: x.stat().st_mtime)

    # 2) читаем Excel
    wb = openpyxl.load_workbook(xls_path, data_only=True)
    ws = wb.active  # в твоем файле лист один: TDSheet

    # ВАЖНО:
    # "пустая строка" проверяем по B (колонка 2), потому что A в данных пустая
    key_col = 2
    read_cols = (2, 3, 6)  # B, C, F -> номенклатура, карриер(код), производитель

    result: Dict[int, List[Any]] = {}
    idx = 1
    empty_streak = 0
    row = 11

    while True:
        key = ws.cell(row=row, column=key_col).value

        # ключ пустой -> считаем пустую строку
        if key is None or (isinstance(key, str) and key.strip() == ""):
            empty_streak += 1
            if empty_streak >= 3:
                break
            row += 1
            continue

        empty_streak = 0

        values = []
        for c in read_cols:
            v = ws.cell(row=row, column=c).value
            v = v.strip() if isinstance(v, str) else v
            values.append(v)

        result[idx] = values
        idx += 1
        row += 1

    wb.close()

    # # 3) переносим файл в ComplenXLS и удаляем из ExcelSource (move)
    # today = datetime.now().strftime("%Y-%m-%d")
    # new_name = f"{today}_session_{session_id}{xls_path.suffix}"
    # dest_path = done_dir / new_name
    # shutil.move(str(xls_path), dest_path)

    return result




if __name__ == "__main__":
    session_id = 1130
    data = load_session_from_excel(session_id)

    print(f"\nРезультат парсинга Excel для session_id={session_id}:\n")
    for k, v in data.items():
        print(f"{k}: {v}")